"""
report.py
=========
Populate report_template.xlsx from the computed figures.

The template ships with the Section/Metric rows already laid out; we fill only
the computed columns (Value, Limit, Utilization, Status, Source) by matching on
the Metric label, preserving the template's own formatting. The Source column
carries the compact graph-path + citation so the report itself is traceable.
"""

from __future__ import annotations

from openpyxl import load_workbook

from .methods import Figure

# Column map (1-indexed) of the template.
COL = {"value": 3, "limit": 4, "utilization": 5, "status": 6, "source": 7}


def write_report(template_path: str, out_path: str, figures: list[Figure]) -> str:
    by_label = {f.label: f for f in figures}
    wb = load_workbook(template_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        metric = row[1].value  # column B
        if not metric:
            continue
        fig = by_label.get(str(metric).strip())
        if fig is None:
            continue
        cit = fig.citation or {}
        source = (f"{fig.graph_path}  |  {cit.get('source_doc')} "
                  f"p.{cit.get('page')} #{cit.get('chunk_id')}")
        ws.cell(row=row[0].row, column=COL["value"]).value = fig.value
        ws.cell(row=row[0].row, column=COL["limit"]).value = fig.limit
        ws.cell(row=row[0].row, column=COL["utilization"]).value = fig.utilization_display
        ws.cell(row=row[0].row, column=COL["status"]).value = fig.status
        ws.cell(row=row[0].row, column=COL["source"]).value = source

    wb.save(out_path)
    return out_path
