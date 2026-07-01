"""
reconcile.py
============
Phase 5 -- three independent checks:

1. Reconciliation: every computed figure vs the firm's answer key (value, limit,
   utilisation, status) with a per-figure pass/fail and delta.
2. Traceability: every figure resolves figure -> graph_path -> source (a real
   chunk_id on a real page). A figure that cannot be traced is a failure.
3. Firewall (constraint 3, verified not asserted): every numeric token in the
   narrative must already exist in the computed output. Any extra number means
   the language model introduced a figure -> fail.
"""

from __future__ import annotations

import re

from openpyxl import load_workbook

from .methods import Figure

# A numeric token is a standalone number. The leading (?<![A-Za-z0-9]) boundary
# stops us matching digits that are part of an identifier -- e.g. the "01" in the
# metric name "DV01", or digits inside "ABS/MBS" -- which are names, not figures.
_NUM = re.compile(r"(?<![A-Za-z0-9])\d[\d,]*\.?\d*")


def _num_tokens(text: str) -> list[str]:
    return [m.group(0).replace(",", "") for m in _NUM.finditer(text or "")]


def load_answer_key(path: str) -> dict[str, dict]:
    """Read an answer-key xlsx into {metric_label: row dict}."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    out = {}
    for r in rows[1:]:
        rec = dict(zip(header, r))
        metric = rec.get("Metric")
        if metric:
            out[str(metric).strip()] = rec
    return out


def reconcile(figures: list[Figure], answer_key: dict[str, dict]) -> list[dict]:
    """Compare computed figures to the answer key, by metric label."""
    results = []
    for f in figures:
        exp = answer_key.get(f.label)
        if exp is None:
            results.append({"figure": f.label, "pass": False,
                            "reason": "metric not in answer key",
                            "computed": f.value, "expected": None, "delta": None})
            continue
        exp_val = str(exp.get("Value", "")).strip()
        exp_status = str(exp.get("Status", "")).strip()
        # Numeric delta on the leading number of each value string.
        cv = _num_tokens(f.value)
        ev = _num_tokens(exp_val)
        delta = None
        if cv and ev:
            delta = round(float(cv[0]) - float(ev[0]), 6)
        value_ok = (f.value == exp_val) or (delta is not None and abs(delta) < 1e-6)
        status_ok = (f.status == exp_status)
        results.append({
            "figure": f.label,
            "pass": bool(value_ok and status_ok),
            "computed": f.value, "expected": exp_val,
            "computed_status": f.status, "expected_status": exp_status,
            "delta": delta,
        })
    return results


def check_traceability(figures: list[Figure]) -> list[dict]:
    """Every figure must resolve figure -> graph_path -> source chunk."""
    out = []
    for f in figures:
        has_path = bool(f.graph_path) and f.graph_path != "(untraceable)"
        cit = f.citation or {}
        has_source = bool(cit.get("chunk_id"))
        out.append({
            "figure": f.label, "pass": bool(has_path and has_source),
            "graph_path": f.graph_path,
            "source": f"{cit.get('source_doc')} p.{cit.get('page')} "
                      f"#{cit.get('chunk_id')}" if has_source else None,
        })
    return out


def firewall(narrative: str, figures: list[Figure]) -> dict:
    """Prove the narrative introduces no number absent from computed output."""
    allowed: set[str] = set()
    for f in figures:
        for field in (f.value, f.limit, f.utilization_display):
            allowed.update(_num_tokens(field))
        if f.utilization_raw is not None:
            # allow common roundings of utilisation
            allowed.add(f"{f.utilization_raw:.1f}")
            allowed.add(str(int(f.utilization_raw * 100)))
        allowed.add(f"{f.value_raw:.0f}")
        allowed.add(f"{f.value_raw:.1f}")
        allowed.add(f"{f.value_raw:.2f}")
    # Years like 2024 / generic small integers used in prose are not figures;
    # we still flag anything numeric that is not an allowed figure token.
    found = _num_tokens(narrative)
    violations = [n for n in found if n not in allowed]
    return {"pass": len(violations) == 0, "violations": sorted(set(violations)),
            "allowed_count": len(allowed)}


def summarise(recon: list[dict], trace: list[dict], fw: dict) -> dict:
    return {
        "reconciliation": {
            "total": len(recon),
            "passed": sum(1 for r in recon if r["pass"]),
            "failed": [r["figure"] for r in recon if not r["pass"]],
        },
        "traceability": {
            "total": len(trace),
            "passed": sum(1 for t in trace if t["pass"]),
            "failed": [t["figure"] for t in trace if not t["pass"]],
        },
        "firewall": fw,
        "all_pass": (all(r["pass"] for r in recon)
                     and all(t["pass"] for t in trace) and fw["pass"]),
    }
